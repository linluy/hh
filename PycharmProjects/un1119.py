# -*- coding: utf-8 -*-


import pandas as pd
import xlwt
import csv
from openpyxl import load_workbook

def getData():
    pd.set_option('display.width',1000)
    pd.set_option('display.max_row',1000)
    pd.set_option('display.max_column',1000)
 
    html=pd.read_html('https://kjc.lnut.edu.cn/info/14022/181217.htm')
    
    bb=pd.ExcelWriter('lgd2019.xlsx')
    html[0].to_excel(bb)
    
    bb.close()
    
    wb = load_workbook('lgd2019.xlsx')
    print(wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
 
    # 调整列宽
    ws.column_dimensions['A'].width = 20.0
    ws.column_dimensions['B'].width = 20.0
    ws.column_dimensions['C'].width = 40.0
    ws.column_dimensions['D'].width = 20.0
 
 
    wb.save('lgd2019.xlsx')

if __name__=="__main__":
    getData()
   
    
    
    
